import csv
import json
import re
import uuid
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Optional

from sqlalchemy import delete, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from models.asin_business_profile import (
    AdEntityAsinMap,
    AdProductDaily,
    AdSearchTermDaily,
    AdTargetDaily,
    AsinBusinessProfile,
    AsinDailySnapshot,
    AsinExecutionLog,
    AsinSkuMap,
    ReportRowStaging,
    ReportUpload,
    ValidationTask,
)
from services.asin_business_profile import DEFAULT_STORE_ID, normalize_asin, normalize_marketplace, parse_float, parse_int


UPLOAD_ROOT = Path("uploads/reports")

REPORT_TYPES = {
    "BUSINESS_REPORT",
    "ADVERTISED_PRODUCT_REPORT",
    "SEARCH_TERM_REPORT",
    "TARGETING_REPORT",
    "BULK_OPERATIONS",
    "PLACEMENT_REPORT",
}

DIRECT_MATCHED = "Direct Matched"
SKU_MATCHED = "SKU Matched"
CAMPAIGN_MAPPED = "Campaign Mapped"
AMBIGUOUS = "Ambiguous"
UNRESOLVED = "Unresolved"

FIELD_ALIASES = {
    "asin": ["asin", "advertised asin", "advertised_asin", "child asin", "child_asin"],
    "sku": ["sku", "seller sku", "seller_sku", "advertised sku", "advertised_sku", "merchant sku"],
    "date": ["date", "start date", "start_date", "report date", "report_date"],
    "product_name": ["product name", "product_name", "title", "item name", "item_name"],
    "brand": ["brand"],
    "category": ["category"],
    "price": ["price", "current price", "current_price"],
    "sessions": ["sessions", "session", "browser sessions", "browser_sessions"],
    "page_views": ["page views", "page_views"],
    "units_ordered": ["units ordered", "units_ordered", "units"],
    "orders": ["orders", "order count", "purchases"],
    "total_sales": ["total sales", "ordered product sales", "sales", "total_sales"],
    "ad_sales": ["7 day total sales", "14 day total sales", "sales", "ad sales", "ad_sales"],
    "ad_spend": ["spend", "ad spend", "ad_spend", "cost"],
    "organic_sales": ["organic sales", "organic_sales"],
    "inventory": ["inventory", "available", "quantity"],
    "buybox_status": ["buybox status", "buybox_status", "buy box status"],
    "impressions": ["impressions"],
    "clicks": ["clicks"],
    "campaign_id": ["campaign id", "campaign_id"],
    "campaign_name": ["campaign name", "campaign_name"],
    "ad_group_id": ["ad group id", "ad_group_id"],
    "ad_group_name": ["ad group name", "ad_group_name"],
    "ad_id": ["ad id", "ad_id"],
    "keyword": ["keyword", "targeting text"],
    "match_type": ["match type", "match_type"],
    "customer_search_term": ["customer search term", "search term", "customer_search_term"],
    "targeting": ["targeting", "target", "targeting expression"],
    "targeting_type": ["targeting type", "targeting_type"],
    "validation_id": ["validation id", "validation_id"],
    "action_type": ["action type", "operation", "entity"],
    "before_value": ["before value", "before_value"],
    "after_value": ["after value", "after_value"],
    "executed_at": ["executed at", "executed_at", "date/time", "datetime"],
    "note": ["note", "notes"],
}


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower().replace("-", " ").replace("_", " "))


def _first(row: dict[str, Any], key: str) -> Any:
    aliases = FIELD_ALIASES.get(key, [key])
    normalized = {_normalize_header(k): v for k, v in row.items()}
    for alias in aliases:
        if _normalize_header(alias) in normalized:
            return normalized[_normalize_header(alias)]
    return None


def _text(value: Any) -> str:
    return str(value or "").strip()


def _parse_date(value: Any, fallback: Optional[date]) -> Optional[date]:
    if isinstance(value, date):
        return value
    text = _text(value)
    if not text:
        return fallback
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(text.split(" ")[0], fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date()
    except ValueError:
        return fallback


def _ratio(numerator: Optional[float], denominator: Optional[float]) -> Optional[float]:
    if numerator is None or denominator in (None, 0):
        return None
    return round(numerator / denominator, 6)


def _read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _read_xlsx(path: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(item or "").strip() for item in rows[0]]
    result = []
    for row in rows[1:]:
        result.append({headers[index]: row[index] if index < len(row) else None for index in range(len(headers))})
    return result


def read_report_rows(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx(path)
    raise ValueError("仅支持 CSV / XLSX")


class ReportImportService:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def create_upload(
        self,
        *,
        seller_id: str,
        store_id: str,
        marketplace: str,
        report_type: str,
        date_range_start: Optional[date],
        date_range_end: Optional[date],
        original_filename: str,
        content: bytes,
        uploaded_by: str,
    ) -> ReportUpload:
        normalized_type = str(report_type or "").strip().upper()
        if normalized_type not in REPORT_TYPES:
            raise ValueError("report_type")
        suffix = Path(original_filename or "").suffix.lower()
        if suffix not in {".csv", ".xlsx", ".xlsm"}:
            raise ValueError("file")

        report_id = f"report_{uuid.uuid4().hex}"
        UPLOAD_ROOT.mkdir(parents=True, exist_ok=True)
        file_path = UPLOAD_ROOT / f"{report_id}{suffix}"
        file_path.write_bytes(content)

        report = ReportUpload(
            report_id=report_id,
            seller_id=seller_id,
            store_id=store_id or DEFAULT_STORE_ID,
            marketplace=normalize_marketplace(marketplace),
            report_type=normalized_type,
            original_filename=original_filename,
            file_path=str(file_path),
            upload_time=datetime.now(timezone.utc),
            uploaded_by=uploaded_by,
            parse_status="Pending",
            date_range_start=date_range_start,
            date_range_end=date_range_end,
        )
        self.db.add(report)
        await self.db.commit()
        await self.db.refresh(report)
        return report

    async def parse_report(self, *, seller_id: str, report_id: str) -> dict[str, Any]:
        result = await self.db.execute(
            select(ReportUpload).where(ReportUpload.report_id == report_id, ReportUpload.seller_id == seller_id)
        )
        report = result.scalar_one_or_none()
        if not report:
            raise ValueError("report_id")
        if not report.file_path:
            raise ValueError("file")

        await self._clear_report_rows(report_id)
        summary = {
            "report_id": report.report_id,
            "report_type": report.report_type,
            "parse_status": "Parsed",
            "total_rows": 0,
            "matched_asin_rows": 0,
            "unmatched_rows": 0,
            "ambiguous_rows": 0,
            "writable_rows": 0,
        }
        try:
            rows = read_report_rows(Path(report.file_path))
            summary["total_rows"] = len(rows)
            for row_number, raw_row in enumerate(rows, start=2):
                normalized = self._normalize_row(raw_row, report)
                match_status, asin = await self._match_asin(report, normalized)
                if match_status in {DIRECT_MATCHED, SKU_MATCHED, CAMPAIGN_MAPPED} and asin and normalized.get("date"):
                    summary["matched_asin_rows"] += 1
                    summary["writable_rows"] += 1
                    normalized["asin"] = asin
                    await self._write_matched_row(report, normalized)
                elif match_status == AMBIGUOUS:
                    summary["ambiguous_rows"] += 1
                else:
                    summary["unmatched_rows"] += 1

                await self._stage_row(
                    report=report,
                    row_number=row_number,
                    raw_row=raw_row,
                    normalized=normalized,
                    match_status=match_status,
                    asin=asin,
                )

            report.parse_status = "Parsed"
            report.parse_error = None
            report.row_count = summary["total_rows"]
            report.matched_rows = summary["matched_asin_rows"]
            report.unresolved_rows = summary["unmatched_rows"]
            report.ambiguous_rows = summary["ambiguous_rows"]
            report.writable_rows = summary["writable_rows"]
            report.match_summary = json.dumps(summary, ensure_ascii=False)
        except Exception as exc:
            report.parse_status = "Failed"
            report.parse_error = str(exc)
            summary["parse_status"] = "Failed"
            report.match_summary = json.dumps(summary, ensure_ascii=False)
        await self.db.commit()
        return summary

    async def get_match_summary(self, *, seller_id: str, report_id: str) -> dict[str, Any]:
        result = await self.db.execute(
            select(ReportUpload).where(ReportUpload.report_id == report_id, ReportUpload.seller_id == seller_id)
        )
        report = result.scalar_one_or_none()
        if not report:
            raise ValueError("report_id")
        if report.match_summary:
            try:
                parsed = json.loads(report.match_summary)
                if isinstance(parsed, dict):
                    return parsed
            except json.JSONDecodeError:
                pass
        return {
            "report_id": report.report_id,
            "report_type": report.report_type,
            "parse_status": report.parse_status,
            "total_rows": report.row_count or 0,
            "matched_asin_rows": report.matched_rows or 0,
            "unmatched_rows": report.unresolved_rows or 0,
            "ambiguous_rows": report.ambiguous_rows or 0,
            "writable_rows": report.writable_rows or 0,
        }

    async def list_staging_rows(
        self,
        *,
        seller_id: str,
        report_id: Optional[str],
        match_status: Optional[str],
        skip: int,
        limit: int,
    ) -> dict[str, Any]:
        query = select(ReportRowStaging).where(ReportRowStaging.seller_id == seller_id)
        count_query = select(func.count(ReportRowStaging.id)).where(ReportRowStaging.seller_id == seller_id)
        if report_id:
            query = query.where(ReportRowStaging.report_id == report_id)
            count_query = count_query.where(ReportRowStaging.report_id == report_id)
        if match_status:
            query = query.where(ReportRowStaging.match_status == match_status)
            count_query = count_query.where(ReportRowStaging.match_status == match_status)
        total_result = await self.db.execute(count_query)
        total = total_result.scalar() or 0
        rows = await self.db.execute(query.order_by(ReportRowStaging.id.desc()).offset(skip).limit(limit))
        return {"items": list(rows.scalars().all()), "total": total, "skip": skip, "limit": limit}

    async def _clear_report_rows(self, report_id: str) -> None:
        for model in (ReportRowStaging, AdProductDaily, AdSearchTermDaily, AdTargetDaily):
            await self.db.execute(delete(model).where(model.source_report_id == report_id) if hasattr(model, "source_report_id") else delete(model).where(model.report_id == report_id))

    def _normalize_row(self, row: dict[str, Any], report: ReportUpload) -> dict[str, Any]:
        spend = parse_float(_first(row, "ad_spend"))
        sales = parse_float(_first(row, "ad_sales")) if report.report_type != "BUSINESS_REPORT" else parse_float(_first(row, "total_sales"))
        clicks = parse_int(_first(row, "clicks"))
        impressions = parse_int(_first(row, "impressions"))
        orders = parse_int(_first(row, "orders"))
        sessions = parse_int(_first(row, "sessions"))
        total_sales = parse_float(_first(row, "total_sales"))
        ad_sales = parse_float(_first(row, "ad_sales"))
        normalized = {
            "asin": normalize_asin(_first(row, "asin")),
            "sku": _text(_first(row, "sku")),
            "date": _parse_date(_first(row, "date"), report.date_range_start),
            "product_name": _text(_first(row, "product_name")),
            "brand": _text(_first(row, "brand")),
            "category": _text(_first(row, "category")),
            "price": parse_float(_first(row, "price")),
            "sessions": sessions,
            "page_views": parse_int(_first(row, "page_views")),
            "units_ordered": parse_int(_first(row, "units_ordered")),
            "orders": orders,
            "total_sales": total_sales,
            "ad_sales": ad_sales,
            "organic_sales": parse_float(_first(row, "organic_sales")),
            "ad_spend": spend,
            "inventory": parse_int(_first(row, "inventory")),
            "buybox_status": _text(_first(row, "buybox_status")),
            "impressions": impressions,
            "clicks": clicks,
            "campaign_id": _text(_first(row, "campaign_id")),
            "campaign_name": _text(_first(row, "campaign_name")),
            "ad_group_id": _text(_first(row, "ad_group_id")),
            "ad_group_name": _text(_first(row, "ad_group_name")),
            "ad_id": _text(_first(row, "ad_id")),
            "keyword": _text(_first(row, "keyword")),
            "match_type": _text(_first(row, "match_type")),
            "customer_search_term": _text(_first(row, "customer_search_term")),
            "targeting": _text(_first(row, "targeting")),
            "targeting_type": _text(_first(row, "targeting_type")),
            "validation_id": _text(_first(row, "validation_id")),
            "action_type": _text(_first(row, "action_type")),
            "before_value": _text(_first(row, "before_value")),
            "after_value": _text(_first(row, "after_value")),
            "executed_at": _parse_date(_first(row, "executed_at"), report.date_range_start),
            "note": _text(_first(row, "note")),
        }
        normalized["ctr"] = _ratio(clicks, impressions)
        normalized["cvr"] = _ratio(orders, sessions)
        normalized["cpc"] = _ratio(spend, clicks)
        normalized["acos"] = _ratio(spend, sales)
        normalized["roas"] = _ratio(sales, spend)
        normalized["tacos"] = _ratio(spend, total_sales)
        return normalized

    async def _match_asin(self, report: ReportUpload, normalized: dict[str, Any]) -> tuple[str, str]:
        asin = normalize_asin(normalized.get("asin"))
        if asin:
            return DIRECT_MATCHED, asin

        sku = _text(normalized.get("sku"))
        if sku:
            sku_matches = await self._match_sku(report, sku)
            if len(sku_matches) == 1:
                return SKU_MATCHED, sku_matches[0]
            if len(sku_matches) > 1:
                return AMBIGUOUS, ""

        entity_matches = await self._match_ad_entity(report, normalized)
        if len(entity_matches) == 1:
            return CAMPAIGN_MAPPED, entity_matches[0]
        if len(entity_matches) > 1:
            return AMBIGUOUS, ""

        if normalized.get("campaign_name") or normalized.get("ad_group_name"):
            return AMBIGUOUS, ""
        return UNRESOLVED, ""

    async def _match_sku(self, report: ReportUpload, sku: str) -> list[str]:
        result = await self.db.execute(
            select(AsinSkuMap.asin)
            .where(
                AsinSkuMap.seller_id == report.seller_id,
                AsinSkuMap.store_id == report.store_id,
                AsinSkuMap.marketplace == normalize_marketplace(report.marketplace),
                or_(AsinSkuMap.sku == sku, AsinSkuMap.seller_sku == sku),
            )
            .distinct()
        )
        return [row[0] for row in result.all() if normalize_asin(row[0])]

    async def _match_ad_entity(self, report: ReportUpload, normalized: dict[str, Any]) -> list[str]:
        ad_id = _text(normalized.get("ad_id"))
        sku = _text(normalized.get("sku"))
        campaign_id = _text(normalized.get("campaign_id"))
        ad_group_id = _text(normalized.get("ad_group_id"))
        campaign_name = _text(normalized.get("campaign_name"))
        ad_group_name = _text(normalized.get("ad_group_name"))
        if not ad_id and not sku:
            return []
        query = select(AdEntityAsinMap.asin).where(
            AdEntityAsinMap.seller_id == report.seller_id,
            AdEntityAsinMap.store_id == report.store_id,
            AdEntityAsinMap.marketplace == normalize_marketplace(report.marketplace),
        )
        if ad_id:
            query = query.where(AdEntityAsinMap.ad_id == ad_id)
            if campaign_id:
                query = query.where(AdEntityAsinMap.campaign_id == campaign_id)
            if ad_group_id:
                query = query.where(AdEntityAsinMap.ad_group_id == ad_group_id)
        elif sku:
            query = query.where(AdEntityAsinMap.sku == sku)
            if campaign_id:
                query = query.where(AdEntityAsinMap.campaign_id == campaign_id)
            elif campaign_name:
                query = query.where(AdEntityAsinMap.campaign_name == campaign_name)
            if ad_group_id:
                query = query.where(AdEntityAsinMap.ad_group_id == ad_group_id)
            elif ad_group_name:
                query = query.where(AdEntityAsinMap.ad_group_name == ad_group_name)
        result = await self.db.execute(query.distinct())
        return [row[0] for row in result.all() if normalize_asin(row[0])]

    async def _write_matched_row(self, report: ReportUpload, normalized: dict[str, Any]) -> None:
        if report.report_type == "BUSINESS_REPORT":
            await self._write_business_report(report, normalized)
        elif report.report_type == "ADVERTISED_PRODUCT_REPORT":
            await self._write_ad_product(report, normalized)
        elif report.report_type == "SEARCH_TERM_REPORT":
            await self._write_search_term(report, normalized)
        elif report.report_type == "TARGETING_REPORT":
            await self._write_target(report, normalized)
        elif report.report_type == "BULK_OPERATIONS":
            await self._write_bulk_mapping(report, normalized)

    async def _write_profile_and_sku(self, report: ReportUpload, normalized: dict[str, Any]) -> None:
        asin = normalize_asin(normalized.get("asin"))
        if not asin:
            return
        result = await self.db.execute(
            select(AsinBusinessProfile).where(
                AsinBusinessProfile.seller_id == report.seller_id,
                AsinBusinessProfile.store_id == report.store_id,
                AsinBusinessProfile.marketplace == normalize_marketplace(report.marketplace),
                AsinBusinessProfile.asin == asin,
            )
        )
        profile = result.scalar_one_or_none()
        if not profile:
            profile = AsinBusinessProfile(
                seller_id=report.seller_id,
                store_id=report.store_id,
                marketplace=normalize_marketplace(report.marketplace),
                asin=asin,
            )
            self.db.add(profile)
        for source, target in (
            ("sku", "sku"),
            ("brand", "brand"),
            ("product_name", "product_name"),
            ("category", "category"),
            ("price", "current_price"),
        ):
            value = normalized.get(source)
            if value not in (None, ""):
                setattr(profile, target, value)
        profile.updated_at = datetime.now(timezone.utc)

        sku = _text(normalized.get("sku"))
        if sku:
            result = await self.db.execute(
                select(AsinSkuMap).where(
                    AsinSkuMap.seller_id == report.seller_id,
                    AsinSkuMap.store_id == report.store_id,
                    AsinSkuMap.marketplace == normalize_marketplace(report.marketplace),
                    AsinSkuMap.sku == sku,
                )
            )
            sku_map = result.scalar_one_or_none()
            if not sku_map:
                sku_map = AsinSkuMap(
                    seller_id=report.seller_id,
                    store_id=report.store_id,
                    marketplace=normalize_marketplace(report.marketplace),
                    sku=sku,
                )
                self.db.add(sku_map)
            sku_map.asin = asin
            sku_map.seller_sku = sku
            sku_map.product_name = normalized.get("product_name") or sku_map.product_name
            sku_map.status = "Direct Matched"

    async def _write_business_report(self, report: ReportUpload, normalized: dict[str, Any]) -> None:
        await self._write_profile_and_sku(report, normalized)
        snapshot = await self._snapshot(report, normalized)
        for key in ("sessions", "page_views", "units_ordered", "orders", "total_sales", "inventory", "buybox_status"):
            if normalized.get(key) not in (None, ""):
                setattr(snapshot, key, normalized.get(key))
        snapshot.sales = normalized.get("total_sales") if normalized.get("total_sales") is not None else snapshot.sales
        snapshot.ctr = normalized.get("ctr") if normalized.get("ctr") is not None else snapshot.ctr
        snapshot.cvr = normalized.get("cvr") if normalized.get("cvr") is not None else snapshot.cvr
        snapshot.source_report_id = report.report_id
        snapshot.data_source = report.report_type

    async def _write_ad_product(self, report: ReportUpload, normalized: dict[str, Any]) -> None:
        await self._write_profile_and_sku(report, normalized)
        self.db.add(
            AdProductDaily(
                seller_id=report.seller_id,
                store_id=report.store_id,
                marketplace=normalize_marketplace(report.marketplace),
                asin=normalize_asin(normalized.get("asin")),
                sku=normalized.get("sku"),
                date=normalized.get("date"),
                period_start=report.date_range_start or normalized.get("date"),
                period_end=report.date_range_end or normalized.get("date"),
                campaign_name=normalized.get("campaign_name"),
                campaign_id=normalized.get("campaign_id"),
                ad_group_name=normalized.get("ad_group_name"),
                ad_group_id=normalized.get("ad_group_id"),
                impressions=normalized.get("impressions"),
                clicks=normalized.get("clicks"),
                spend=normalized.get("ad_spend"),
                sales=normalized.get("ad_sales"),
                orders=normalized.get("orders"),
                units=normalized.get("units_ordered"),
                ctr=normalized.get("ctr"),
                cpc=normalized.get("cpc"),
                acos=normalized.get("acos"),
                roas=normalized.get("roas"),
                source_report_id=report.report_id,
            )
        )
        snapshot = await self._snapshot(report, normalized)
        snapshot.impressions = (snapshot.impressions or 0) + (normalized.get("impressions") or 0)
        snapshot.clicks = (snapshot.clicks or 0) + (normalized.get("clicks") or 0)
        snapshot.ad_spend = (snapshot.ad_spend or 0) + (normalized.get("ad_spend") or 0)
        snapshot.ad_sales = (snapshot.ad_sales or 0) + (normalized.get("ad_sales") or 0)
        snapshot.acos = _ratio(snapshot.ad_spend, snapshot.ad_sales)
        snapshot.tacos = _ratio(snapshot.ad_spend, snapshot.total_sales)
        snapshot.source_report_id = report.report_id
        await self._upsert_ad_entity_map(report, normalized)

    async def _write_search_term(self, report: ReportUpload, normalized: dict[str, Any]) -> None:
        self.db.add(
            AdSearchTermDaily(
                seller_id=report.seller_id,
                store_id=report.store_id,
                marketplace=normalize_marketplace(report.marketplace),
                asin=normalize_asin(normalized.get("asin")),
                date=normalized.get("date"),
                period_start=report.date_range_start or normalized.get("date"),
                period_end=report.date_range_end or normalized.get("date"),
                campaign_name=normalized.get("campaign_name"),
                ad_group_name=normalized.get("ad_group_name"),
                keyword=normalized.get("keyword"),
                match_type=normalized.get("match_type"),
                customer_search_term=normalized.get("customer_search_term"),
                impressions=normalized.get("impressions"),
                clicks=normalized.get("clicks"),
                spend=normalized.get("ad_spend"),
                sales=normalized.get("ad_sales"),
                orders=normalized.get("orders"),
                units=normalized.get("units_ordered"),
                ctr=normalized.get("ctr"),
                cpc=normalized.get("cpc"),
                acos=normalized.get("acos"),
                roas=normalized.get("roas"),
                source_report_id=report.report_id,
            )
        )

    async def _write_target(self, report: ReportUpload, normalized: dict[str, Any]) -> None:
        self.db.add(
            AdTargetDaily(
                seller_id=report.seller_id,
                store_id=report.store_id,
                marketplace=normalize_marketplace(report.marketplace),
                asin=normalize_asin(normalized.get("asin")),
                date=normalized.get("date"),
                campaign_name=normalized.get("campaign_name"),
                ad_group_name=normalized.get("ad_group_name"),
                targeting=normalized.get("targeting"),
                targeting_type=normalized.get("targeting_type"),
                match_type=normalized.get("match_type"),
                impressions=normalized.get("impressions"),
                clicks=normalized.get("clicks"),
                spend=normalized.get("ad_spend"),
                sales=normalized.get("ad_sales"),
                orders=normalized.get("orders"),
                units=normalized.get("units_ordered"),
                ctr=normalized.get("ctr"),
                cpc=normalized.get("cpc"),
                acos=normalized.get("acos"),
                roas=normalized.get("roas"),
                source_report_id=report.report_id,
            )
        )

    async def _write_bulk_mapping(self, report: ReportUpload, normalized: dict[str, Any]) -> None:
        await self._write_profile_and_sku(report, normalized)
        await self._upsert_ad_entity_map(report, normalized)
        await self._write_bulk_execution_log(report, normalized)

    async def _write_bulk_execution_log(self, report: ReportUpload, normalized: dict[str, Any]) -> None:
        validation_id = _text(normalized.get("validation_id"))
        action_type = _text(normalized.get("action_type"))
        asin = normalize_asin(normalized.get("asin"))
        if not validation_id or not action_type or not asin:
            return
        result = await self.db.execute(
            select(ValidationTask.validation_id).where(
                ValidationTask.validation_id == validation_id,
                ValidationTask.seller_id == report.seller_id,
                ValidationTask.store_id == report.store_id,
                ValidationTask.marketplace == normalize_marketplace(report.marketplace),
                ValidationTask.asin == asin,
            )
        )
        if not result.scalar_one_or_none():
            return
        self.db.add(
            AsinExecutionLog(
                execution_id=f"exec_{uuid.uuid4().hex}",
                validation_id=validation_id,
                seller_id=report.seller_id,
                store_id=report.store_id,
                marketplace=normalize_marketplace(report.marketplace),
                asin=asin,
                action_type=action_type,
                before_value=normalized.get("before_value"),
                after_value=normalized.get("after_value"),
                executed_by=report.uploaded_by or report.seller_id,
                executed_at=datetime.combine(normalized.get("executed_at"), datetime.min.time(), tzinfo=timezone.utc)
                if normalized.get("executed_at")
                else datetime.now(timezone.utc),
                note=normalized.get("note"),
                source="BULK_OPERATIONS",
                data_source=report.report_type,
            )
        )

    async def _snapshot(self, report: ReportUpload, normalized: dict[str, Any]) -> AsinDailySnapshot:
        result = await self.db.execute(
            select(AsinDailySnapshot).where(
                AsinDailySnapshot.seller_id == report.seller_id,
                AsinDailySnapshot.store_id == report.store_id,
                AsinDailySnapshot.marketplace == normalize_marketplace(report.marketplace),
                AsinDailySnapshot.asin == normalize_asin(normalized.get("asin")),
                AsinDailySnapshot.date == normalized.get("date"),
            )
        )
        snapshot = result.scalar_one_or_none()
        if not snapshot:
            snapshot = AsinDailySnapshot(
                seller_id=report.seller_id,
                store_id=report.store_id,
                marketplace=normalize_marketplace(report.marketplace),
                asin=normalize_asin(normalized.get("asin")),
                date=normalized.get("date"),
            )
            self.db.add(snapshot)
        return snapshot

    async def resolve_staging_rows(
        self,
        *,
        seller_id: str,
        report_id: str,
        action: str,
        asin: Optional[str],
        staging_row_ids: list[int],
    ) -> dict[str, Any]:
        normalized_action = str(action or "").strip()
        if normalized_action not in {"bind_existing", "create_profile", "ignore"}:
            raise ValueError("action")
        result = await self.db.execute(
            select(ReportUpload).where(ReportUpload.report_id == report_id, ReportUpload.seller_id == seller_id)
        )
        report = result.scalar_one_or_none()
        if not report:
            raise ValueError("report_id")

        normalized_asin = normalize_asin(asin) if asin else ""
        if normalized_action in {"bind_existing", "create_profile"} and not normalized_asin:
            raise ValueError("asin")

        query = select(ReportRowStaging).where(
            ReportRowStaging.report_id == report.report_id,
            ReportRowStaging.seller_id == seller_id,
            ReportRowStaging.match_status.in_([AMBIGUOUS, UNRESOLVED]),
        )
        if staging_row_ids:
            query = query.where(ReportRowStaging.id.in_(staging_row_ids))
        rows = list((await self.db.execute(query.order_by(ReportRowStaging.id.asc()))).scalars().all())

        summary = {
            "report_id": report.report_id,
            "report_type": report.report_type,
            "parse_status": report.parse_status or "Parsed",
            "action": normalized_action,
            "total_rows": len(rows),
            "matched_asin_rows": 0,
            "unmatched_rows": 0,
            "ambiguous_rows": 0,
            "writable_rows": 0,
        }
        if normalized_action == "ignore":
            for row in rows:
                await self.db.delete(row)
            summary["unmatched_rows"] = len(rows)
            await self.db.commit()
            return summary

        if normalized_action == "create_profile":
            profile = AsinBusinessProfile(
                seller_id=report.seller_id,
                store_id=report.store_id,
                marketplace=normalize_marketplace(report.marketplace),
                asin=normalized_asin,
                data_source="REPORT_ROW_STAGING",
            )
            existing = await self.db.execute(
                select(AsinBusinessProfile).where(
                    AsinBusinessProfile.seller_id == report.seller_id,
                    AsinBusinessProfile.store_id == report.store_id,
                    AsinBusinessProfile.marketplace == normalize_marketplace(report.marketplace),
                    AsinBusinessProfile.asin == normalized_asin,
                )
            )
            if not existing.scalar_one_or_none():
                self.db.add(profile)

        for row in rows:
            normalized = json.loads(row.normalized_data or "{}")
            if not isinstance(normalized, dict):
                normalized = {}
            normalized["asin"] = normalized_asin
            normalized["date"] = _parse_date(normalized.get("date"), report.date_range_start)
            row.asin = normalized_asin
            row.matched_asin = normalized_asin
            row.match_method = "Manual"
            row.is_writable = bool(normalized.get("date"))
            row.resolution_status = "Resolved" if normalized.get("date") else "Pending"
            row.normalized_data = json.dumps(normalized, ensure_ascii=False, default=str)
            if not normalized.get("date"):
                row.match_status = UNRESOLVED
                summary["unmatched_rows"] += 1
                continue
            row.match_status = DIRECT_MATCHED
            await self._write_matched_row(report, normalized)
            summary["matched_asin_rows"] += 1
            summary["writable_rows"] += 1

            report.parse_status = "Parsed"
            report.matched_rows = (report.matched_rows or 0) + summary["matched_asin_rows"]
            report.unresolved_rows = max((report.unresolved_rows or 0) - summary["matched_asin_rows"], 0)
            report.ambiguous_rows = max((report.ambiguous_rows or 0) - summary["matched_asin_rows"], 0)
            report.writable_rows = (report.writable_rows or 0) + summary["writable_rows"]
            report.match_summary = json.dumps(summary, ensure_ascii=False)
        await self.db.commit()
        return summary

    async def _upsert_ad_entity_map(self, report: ReportUpload, normalized: dict[str, Any]) -> None:
        asin = normalize_asin(normalized.get("asin"))
        if not asin:
            return
        ad_id = _text(normalized.get("ad_id"))
        sku = _text(normalized.get("sku"))
        if not ad_id and not sku:
            return
        query = select(AdEntityAsinMap).where(
            AdEntityAsinMap.seller_id == report.seller_id,
            AdEntityAsinMap.store_id == report.store_id,
            AdEntityAsinMap.marketplace == normalize_marketplace(report.marketplace),
        )
        if ad_id:
            query = query.where(AdEntityAsinMap.ad_id == ad_id)
        else:
            query = query.where(AdEntityAsinMap.sku == sku)
        result = await self.db.execute(query)
        mapping = result.scalars().first()
        if not mapping:
            mapping = AdEntityAsinMap(
                seller_id=report.seller_id,
                store_id=report.store_id,
                marketplace=normalize_marketplace(report.marketplace),
            )
            self.db.add(mapping)
        mapping.asin = asin
        mapping.sku = sku or mapping.sku
        mapping.campaign_id = normalized.get("campaign_id") or mapping.campaign_id
        mapping.campaign_name = normalized.get("campaign_name") or mapping.campaign_name
        mapping.ad_group_id = normalized.get("ad_group_id") or mapping.ad_group_id
        mapping.ad_group_name = normalized.get("ad_group_name") or mapping.ad_group_name
        mapping.ad_id = ad_id or mapping.ad_id

    async def _stage_row(
        self,
        *,
        report: ReportUpload,
        row_number: int,
        raw_row: dict[str, Any],
        normalized: dict[str, Any],
        match_status: str,
        asin: str,
    ) -> None:
        self.db.add(
            ReportRowStaging(
                report_id=report.report_id,
                seller_id=report.seller_id,
                store_id=report.store_id,
                marketplace=normalize_marketplace(report.marketplace),
                asin=asin or None,
                date=normalized.get("date"),
                row_number=row_number,
                report_type=report.report_type,
                match_status=match_status,
                match_method=match_status,
                extracted_asin=normalized.get("asin"),
                extracted_sku=normalized.get("sku"),
                campaign_id=normalized.get("campaign_id"),
                ad_group_id=normalized.get("ad_group_id"),
                ad_id=normalized.get("ad_id"),
                matched_asin=asin or None,
                candidate_matches=None,
                is_writable=bool(match_status in {DIRECT_MATCHED, SKU_MATCHED, CAMPAIGN_MAPPED} and asin and normalized.get("date")),
                resolution_status="Resolved" if match_status in {DIRECT_MATCHED, SKU_MATCHED, CAMPAIGN_MAPPED} else "Pending",
                raw_data=json.dumps(raw_row, ensure_ascii=False, default=str),
                normalized_data=json.dumps(normalized, ensure_ascii=False, default=str),
            )
        )
